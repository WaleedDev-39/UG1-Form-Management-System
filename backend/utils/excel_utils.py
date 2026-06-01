from io import BytesIO
from datetime import datetime
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# UAF brand colours
GREEN  = "1A6B3A"
GOLD   = "C9A84C"
LGREY  = "F0F7F2"
WHITE  = "FFFFFF"
RED_L  = "F8D7DA"
YEL_L  = "FFF3CD"
GRN_L  = "D4EDDA"

_thin  = Side(style="thin",   color="CCCCCC")
_med   = Side(style="medium", color="999999")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor=GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = Border(left=_thin, right=_thin, top=_thin, bottom=Side(style="medium", color=WHITE))
    return c


def generate_ug1_report(forms_data: List[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UG1 Report"

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value     = "University of Agriculture Faisalabad – UG-1 Form Management Report"
    tc.font      = Font(bold=True, size=13, color=WHITE)
    tc.fill      = PatternFill("solid", fgColor=GOLD)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    dc = ws["A2"]
    dc.value     = f"Generated: {datetime.now().strftime('%d-%b-%Y  %H:%M')}"
    dc.font      = Font(italic=True, size=9)
    dc.alignment = Alignment(horizontal="center")

    # ── Headers ────────────────────────────────────────────────────────────
    headers = [
        "S.No", "AG Number", "Student Name", "Father's Name",
        "CNIC", "Boarder Status", "Semester",
        "Voucher ID", "Fee Status",
        "Submission Date", "Processed By", "Notes",
    ]
    HDR_ROW = 4
    for col, h in enumerate(headers, 1):
        _hdr_cell(ws, HDR_ROW, col, h)
    ws.row_dimensions[HDR_ROW].height = 22

    # ── Data ───────────────────────────────────────────────────────────────
    status_fill = {
        "Paid":        PatternFill("solid", fgColor=GRN_L),
        "Installment": PatternFill("solid", fgColor=YEL_L),
        "Deferred":    PatternFill("solid", fgColor=RED_L),
    }

    for i, f in enumerate(forms_data, 1):
        r = HDR_ROW + i
        fee_st = f.get("fee_status", "")
        row_fill = status_fill.get(fee_st, PatternFill("solid", fgColor=LGREY if i % 2 == 0 else WHITE))

        values = [
            i,
            f.get("ag_number", ""),
            f.get("student_name", ""),
            f.get("father_name", ""),
            f.get("cnic", ""),
            f.get("boarder_status", ""),
            f.get("semester", ""),
            f.get("voucher_id", ""),
            fee_st,
            f.get("submission_date", ""),
            f.get("processed_by", ""),
            f.get("notes", ""),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = row_fill
            c.border    = _border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col == 9:  # fee status bold
                c.font = Font(bold=True, size=10)
        ws.row_dimensions[r].height = 18

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = [6, 16, 22, 22, 16, 14, 10, 16, 14, 20, 18, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{HDR_ROW + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
